from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Border, Side, Font

workbook_path = input("Ingrese la ruta del archivo de Excel: ").strip('"').replace("\\", "\\\\")


def add_transaction_to_excel(path, data, entry_number):
    try:
        wb = load_workbook(path)
        ws = wb.active

        last_separator_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "**********************************************************":
                last_separator_row = row

        if last_separator_row is None:
            start_row = 4
        else:
            start_row = last_separator_row + 1

        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        font = Font(bold=True)
        for col, value in enumerate(["FECHA", "DETALLE", "AUX", "DÉBITO", "CRÉDITO"], start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = value
            cell.font = font
            cell.border = border

        fecha_cell = ws.cell(row=start_row, column=1, value=data["fecha"].strftime("%d/%m/%Y"))
        fecha_cell.border = border

        ws.cell(row=start_row, column=2, value=data["control_deudor"]).border = border
        ws.cell(row=start_row, column=4, value=float(data["debito"])).border = border

        ws.cell(row=start_row + 1, column=2, value=data["auxiliar_deudor"]).border = border
        ws.cell(row=start_row + 1, column=3, value=float(data["debito"])).border = border

        ws.cell(row=start_row + 2, column=2, value="@").border = border

        ws.cell(row=start_row + 3, column=2, value=data["control_acreedor"]).border = border
        ws.cell(row=start_row + 3, column=5, value=float(data["credito"])).border = border

        ws.cell(row=start_row + 4, column=2, value=data["auxiliar_acreedor"]).border = border
        ws.cell(row=start_row + 4, column=3, value=float(data["credito"])).border = border

        for row in range(start_row, start_row + 5):
            for col in range(3, 6):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0.00'

        input_text = input("P/R: ").upper()

        lines = []
        line = ""
        for word in input_text.split():
            if len(line) + len(word) > 60:
                lines.append(line)
                line = word
            else:
                if line:
                    line += " "
                line += word
        lines.append(line)

        for i, line in enumerate(lines):
            if i == 0:
                ws.cell(row=start_row + 5 + i, column=2, value="P/R " + line).border = border
            else:
                ws.cell(row=start_row + 5 + i, column=2, value=line).border = border

        ws.cell(row=start_row + 5 + len(lines), column=2,
                value=f"E/D # {entry_number}-{data['fecha'].strftime('%m-%Y')}").border = border

        ws.cell(row=start_row + 5 + len(lines) + 1, column=1,
                value="**********************************************************").border = border

        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "**********************************************************":
                last_separator_row = row

        for row in range(start_row, last_separator_row + 1):
            for column in range(1, 6):
                ws.cell(row=row, column=column).border = border

        wb.save(path)
    except Exception as e:
        print(f"Error al agregar la transacción: {e}")


def add_transaction(diccionario):
    control_deudor = input("Inserte el nombre de la primera cuenta control: ").upper()
    auxiliar_deudor = input("Inserte el nombre de la primera cuenta auxiliar: ").upper()

    while True:
        try:
            debito = float(input("Ingrese el monto a debitar: "))
            break
        except ValueError:
            print("Error: Por favor, ingrese un monto válido.")

    control_acreedor = input("Inserte el nombre de la segunda cuenta control: ").upper()
    auxiliar_acreedor = input("Inserte el nombre de la segunda cuenta auxiliar: ").upper()

    while True:
        try:
            fecha = input("Ingrese la fecha de la transacción en formato DD/MM/YYYY: ")
            formats = ["%d/%m/%Y", "%d-%m-%y", "%y %d/%m/%Y", "%d-%m-%Y"]
            date_obj = None
            for fmt in formats:
                try:
                    date_obj = datetime.strptime(fecha, fmt)
                    break
                except ValueError:
                    pass
            if date_obj is None:
                raise ValueError("Formato de fecha inválido. Por favor, ingrese una fecha válida en uno de los "
                                 "formatos especificados.")
            break
        except ValueError as e:
            print(f"Error: {e}")

    while True:
        try:
            credito = float(input("Ingrese el monto a acreditar: "))
            break
        except ValueError:
            print("Error: Por favor, ingrese un monto válido.")

    while True:
        try:
            entry_number = int(input("Ingrese el número de la entrada de diario: "))
            break
        except ValueError:
            print("Error: Por favor, ingrese un número entero válido para el número de la entrada de diario.")

    diccionario["Cuentas"] = {
        "fecha": date_obj,
        "control_deudor": control_deudor,
        "control_acreedor": control_acreedor,
        "auxiliar_deudor": auxiliar_deudor,
        "auxiliar_acreedor": auxiliar_acreedor,
        "debito": debito,
        "credito": credito
    }

    add_transaction_to_excel(workbook_path, diccionario["Cuentas"], entry_number)



mi_diccionario = {}
while True:
    print("1. Crear nueva transacción")
    print("2. Cerrar programa")
    opcion = input("Seleccione una opción: ")

    if opcion == "1":
        add_transaction(mi_diccionario)
    elif opcion == "2":
        break
    else:
        print("Opción inválida. Por favor, seleccione una opción válida.")